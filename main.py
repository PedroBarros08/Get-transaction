import requests
from requests.exceptions import ConnectionError, HTTPError
import pandas as pd
from openpyxl import Workbook
import logging
import os
from dotenv import load_dotenv


load_dotenv(dotenv_path=".env")

client_id = os.getenv("CLIENT_ID")
client_secret = os.getenv("CLIENT_SECRET")
resource = os.getenv("RESOURCE") 
url_get = os.getenv("GET_URL")
url_token = os.getenv("GET_TOKEN")



logging.basicConfig(format="%(asctime)s-%(levelname)s-%(message)s", level=logging.INFO, filename="arquivo.log", encoding="utf-8" )

def autenticacao():
    """
    Esta função permite gerar token usando os seguintes dados:
    grant_type, client secret, client id e resource. 
    """
    try:
        url = f"{url_token}"
        payload = f'grant_type=client_credentials&client_id={client_id}&client_secret={client_secret}&resource={resource}'
        headers = {
        'Content-Type': 'application/x-www-form-urlencoded',
        'Cookie': 'buid=1.ASAA4Uk7UXbS80mADfy5jzEqb23-VQj4EfdMo56LOLmqI-s9AQAgAA.AQABGgEAAABVrSpeuWamRam2jAF1XRQESmtc3zjVSSqapZC5sjDMcr--e24ZZvC9tPwBQdLiuyWKtlElyeNDtLebT4OB_Xw5Hfpap0wmBdSi92_Io7HiOxvotrR2C_5-JAYWmmtnK6IgAA; fpc=AmM23xn0fgVAvD1tXOIiBz5yDpLVAQAAABd8huAOAAAA; stsservicecookie=estsfd; x-ms-gateway-slice=estsfd; fpc=AmM23xn0fgVAvD1tXOIiBz6dzjl4AQAAAM4RoeEOAAAA'
        }

        logging.info("Fazendo a requisição do token...")

        response = requests.get(url, headers=headers, data=payload)

        response.raise_for_status()

        return response
    
    except ConnectionError as conn_err:

        logging.error(f"Ocorreu um erro de conexão: {conn_err}")

    except HTTPError as http_err: 

        logging.error(f'Ocorreu um erro HTTP: {http_err}')

    except Exception as e:

        logging.error(f"Ocorreu um erro inesperado: {e}")

    else:
        logging.info(f"Token gerado com sucesso, status code: {response.status_code}")
    

    

def extract():
    """
    Extrai aos dados levando em conta a paginação da API.
    """
   
    token = autenticacao().json()["access_token"]
    
    todas_transacoes = []
    skip=0
    contagem = 0

    logging.info("Fazendo requisição à API das transações")

    while True:

        url = f"{url_get}?skip={skip}"
        payload = {}
        headers = {"Authorization": f"Bearer {token}"}
        request = requests.get(url, headers=headers, data=payload)
        response = request.json() 

        if request.status_code == 401:
            logging.warning("Token expirado. Renovando...")
            token_refresh = autenticacao()
            token = token_refresh.json()["access_token"]
            logging.info("Token renovado com sucesso")
            continue


        
        dados = response.get('payments')
        todas_transacoes.extend(dados)

        if not response.get("hasMorePages", False):
            logging.info(f"{skip} - última página da API")
            break
        skip += 50

       
        

    return todas_transacoes

def transaform():
    try:
        transacoes = extract()
        logging.info("Transaformar os dados em DataFrame")
        df = pd.json_normalize(transacoes)
        return  df
    except Exception as e:
        logging.error(f"Ocorreu um erro inesperado {e}")
    else:
        logging.info("Dados transformado com sucesso")
   

def excel():

    df = transaform()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos_sucessos"

    wb.create_sheet("Pagamentos_pendente")
    wb.create_sheet("Pagamentos_expirados")
    wb.create_sheet("Pagamentos_falhado")
    wb.create_sheet("outros_estados")

    sheet_sucesso = wb["Pagamentos_sucessos"]
    sheet_pendente = wb["Pagamentos_pendente"]
    sheet_falhado = wb["Pagamentos_falhado"]
    sheet_expirado = wb["Pagamentos_expirados"]
    sheet_outros_estados = wb["outros_estados"]


    sheet_sucesso.append(["id", "amount", "currency", "status", "paymentMethod", "createdDate", "updatedDate", "dueDate", "entity"])
    sheet_falhado.append(["id", "amount", "currency", "status", "paymentMethod", "createdDate", "updatedDate", "dueDate", "entity"])
    sheet_pendente.append(["id", "amount", "currency", "status", "paymentMethod", "createdDate", "updatedDate", "dueDate", "entity"])
    sheet_expirado.append(["id", "amount", "currency", "status", "paymentMethod", "createdDate", "updatedDate", "dueDate", "entity"])

    for celula in df.itertuples():
        try:
            index, id, merchantTransactionId, type, operation, amount, currency, status, description, disputes, applicationFeeAmount, paymentMethod,  createdDate, updatedDate, options, reference, dueDate, entity, referencia = celula
            if status == "Success":
                sheet_sucesso.append([id, amount, currency, status, paymentMethod, createdDate, updatedDate, dueDate, entity])
            elif status == "Pending":
                sheet_pendente.append([id, amount, currency, status, paymentMethod, createdDate, updatedDate, dueDate, entity])
            elif status == "Expired":
                sheet_expirado.append([id, amount, currency, status, paymentMethod, createdDate, updatedDate, dueDate, entity])
            elif status == "Failed":
                sheet_falhado.append([id, amount, currency, status, paymentMethod, createdDate, updatedDate, dueDate, entity])
            else:
                sheet_outros_estados.append(["id", "amount", "currency", "status", "paymentMethod", "createdDate", "updatedDate", "dueDate", "entity"])
                sheet_outros_estados.append([id, amount, currency, status, paymentMethod, createdDate, updatedDate, dueDate, entity])
        except Exception as e:
            logging.error("Ocorreu um erro inesperado {e} ao escrever no ficheiro excel")
        else:
            logging.info("Dados preenchido no ficheiro excel com sucesso")


    logging.info("Salvar o documento...")
    return wb.save("transacoes.xlsx")
    


    
if __name__ == "__main__":
    excel()
